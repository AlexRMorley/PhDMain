#!/usr/bin/env python3
"""
ScalingBenchL.py — Framework L fleet-size scaling curve.

Sweeps fleet size N ∈ {10..100, 1000} for Framework L ONLY (belief-safe K).
No baseline models — this is a scaling study of L in isolation.

KEY DESIGN — constant robot density:
    The map grows with the fleet so cells-per-robot stays ≈ constant. This
    isolates "does per-step COORDINATION cost scale with N" from "does the
    mission get trivially easy/hard as robots are packed denser". Anchor:
    50 robots on 200×200 = 40000 cells ⇒ 800 cells/robot. Every N uses
    grid = round(sqrt(N × 800)), clamped even, so density is held.
    (Set --fixed-grid to instead hold the map at 200×200 and let density rise
    — useful as a contrasting saturation curve.)

Robot TYPE MIX held at the large-bench proportion:
    Legged 26% / Drone 34% / Boat 16% / Rover 24%  (13:17:8:12 → 50).
Survivors scale with map area at the large-bench density (45 per 40000 cells).

COMPUTE INSTRUMENTATION (per step, summarised over the run):
    step_total     — full sim.step() wall time
    role           — _pg_best_response_roles (potential game)
    alloc          — _assign_zones_cbba (HRBA)
    cover          — relay coverage union + disk machinery
    astar_sum      — Σ over robots of A* time on the single sim CPU
    astar_max      — slowest single robot's A* this step (parallel proxy)
    ground         — step_total − astar_sum  (centralised coordination only)
    dist_latency   — ground + astar_max      (realistic per-step latency if
                                              each robot plans onboard)
  Each reported as mean, p50, p90, max, AND per-robot (mean / N) so O(N)
  vs O(N²) vs O(1) phase scaling is directly readable.

OUTPUTS (to --out dir):
    scaling_L_summary.xlsx  — formatted workbook, one row per (N, seed) on a
                              raw sheet + an aggregated-by-N sheet + a config
                              sheet + a phase-scaling sheet with Excel formulas.
    scaling_L_long.csv      — tidy/long format, one row per (N, seed, metric)
                              — ideal for R (read.csv → ggplot/lm/anova).
    scaling_L_wide.csv      — one row per (N, seed), all metrics as columns.
    scaling_L_curve.png     — quick-look scaling plots.

Usage:
    python3 ScalingBenchL.py --seeds 3 --steps 1500       # seeds 0,1,2
    python3 ScalingBenchL.py --quick                     # tiny smoke
    python3 ScalingBenchL.py --sizes 10,20,50 --steps 800
    python3 ScalingBenchL.py --seeds 1,2,7,42             # explicit seeds
    PYTHONHASHSEED=0 python3 ScalingBenchL.py --seeds 5  # reproducible
"""
import argparse, csv, importlib.util, math, os, random, sys, time
from collections import defaultdict
import numpy as np

# ── Load Framework L ─────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
_SEARCH_DIRS = [_HERE, os.path.join(_HERE, '..'), os.path.join(_HERE, '..', 'outputs'),
                '/mnt/user-data/outputs', '/home/claude']

def _find(*names):
    for name in names:
        for folder in _SEARCH_DIRS:
            p = os.path.join(folder, name)
            if os.path.exists(p):
                return os.path.abspath(p)
    print("\nERROR: Cannot find any of:", list(names))
    print("Searched folders:")
    for d in _SEARCH_DIRS:
        print("  " + d)
    sys.exit(1)

def _load(path, name):
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec); sys.modules[name] = mod
    spec.loader.exec_module(mod); return mod

_FRAMEWORK_PATH = _find('2DFleetFrameworkM.py', 'FleetFrameworkM.py')
_CLUSTER_ANCHOR = "k = max(1, min(8, round(R ** 0.5)))"

def _load_framework(cluster_cap=None):
    """Load the framework, optionally re-writing the HRBA Layer-0 cluster cap.

    The framework caps the spatial cluster count at 8 (k = min(8, sqrt(R))).
    Beyond ~64 robots no new clusters are added, so the hierarchical auction
    stops subdividing exactly in the range where alloc cost bends superlinear
    — the cap is the designed mitigation's own ceiling. cluster_cap:
      None -> framework as shipped (cap 8)
      0    -> UNCAPPED, k = round(sqrt(R))  (the scaling experiment)
      n>=1 -> cap at n
    Applied as an exact-anchor source patch to a temp copy; the framework
    file itself is never modified. Aborts loudly if the anchor is missing."""
    if cluster_cap is None:
        return _load(_FRAMEWORK_PATH, 'FrameworkM')
    s = open(_FRAMEWORK_PATH, 'r', encoding='utf-8', errors='replace').read()
    n = s.count(_CLUSTER_ANCHOR)
    if n != 1:
        print(f"ERROR: cluster-cap anchor found {n}x (expected 1) — framework "
              f"changed; update _CLUSTER_ANCHOR in this bench."); sys.exit(1)
    if cluster_cap == 0:
        repl = "k = max(1, round(R ** 0.5))  # [SCALING BENCH] cluster cap REMOVED"
    else:
        repl = (f"k = max(1, min({cluster_cap}, round(R ** 0.5)))"
                f"  # [SCALING BENCH] cluster cap {cluster_cap}")
    s = s.replace(_CLUSTER_ANCHOR, repl)
    import tempfile
    tmp = os.path.join(tempfile.gettempdir(), f'_scaling_fw_cap{cluster_cap}.py')
    with open(tmp, 'w', encoding='utf-8') as f:
        f.write(s)
    return _load(tmp, f'FrameworkM_cap{cluster_cap}')

L = _load_framework(None)

# ── A* timing wrap (module-global accumulator, reset each step) ───────────────
# _ASTAR is the ALL-CALLS total (unchanged, kept for backward compatibility with
# the existing astar_sum/astar_max/dist_* columns).
# _ASTAR_PHASE splits that same total by CALL SITE, using a context flag set by
# whichever FleetSim method is currently invoking it — see run_one() below,
# which wraps _move_robot (phase='explorer': Robot.set_goal + Robot.move_step's
# replan, both call AStar.search directly) and _move_relay (phase='relay':
# its own AStar.search call). Neither _assign_zones_cbba (alloc) nor
# _pg_best_response_roles (role) call AStar.search at all — A* time was never
# part of 'alloc', it was silently absorbed into the unlabelled 'other'
# residual. This split makes that visible instead of hidden.
_ASTAR = {'sum': 0.0, 'max': 0.0, 'n': 0}
_ASTAR_PHASE = {'explorer': 0.0, 'relay': 0.0, 'other': 0.0}
_CURRENT_PHASE = {'p': None}
def _wrap_astar(cls):
    _orig = cls.search
    def _timed(*a, **k):
        t0 = time.perf_counter(); r = _orig(*a, **k)
        dt = (time.perf_counter() - t0) * 1000.0
        _ASTAR['sum'] += dt
        if dt > _ASTAR['max']: _ASTAR['max'] = dt
        _ASTAR['n'] += 1
        _ASTAR_PHASE[_CURRENT_PHASE['p'] or 'other'] += dt
        return r
    cls.search = staticmethod(_timed)
if hasattr(L, 'AStar') and hasattr(L.AStar, 'search'):
    _wrap_astar(L.AStar)

# ── Scaling configuration ────────────────────────────────────────────────────
DENSITY_CELLS_PER_ROBOT = 800          # 50 robots @ 200×200 anchor
SURV_PER_CELL           = 45 / 40000.0  # large-bench survivor density
TYPE_MIX = [("Legged", 13), ("Drone", 17), ("Boat", 8), ("Rover", 12)]
_MIX_TOTAL = sum(n for _, n in TYPE_MIX)  # 50

def grid_for(n_robots, fixed_grid):
    if fixed_grid:
        return 200
    g = int(round(math.sqrt(n_robots * DENSITY_CELLS_PER_ROBOT)))
    return max(60, g + (g & 1))          # even, min 60

def counts_for(n_robots):
    """Split N into the four types preserving the mix; remainder to Drone."""
    raw = {t: n_robots * n / _MIX_TOTAL for t, n in TYPE_MIX}
    out = {t: int(math.floor(v)) for t, v in raw.items()}
    short = n_robots - sum(out.values())
    for t in sorted(raw, key=lambda k: raw[k] - out[k], reverse=True):
        if short <= 0: break
        out[t] += 1; short -= 1
    return out

# ── Build a Framework L sim at a given scale (patches module constants) ───────
def make_sim(n_robots, grid, n_survivors):
    L.GRID_W = grid; L.GRID_H = grid
    counts = counts_for(n_robots)

    def _build_robots(self):
        Cap = L.Capability
        templates = {
            "Legged": ({Cap.LAND, Cap.STAIRS}, np.array([10., 10.]), (L.TEMP_LIMIT, L.RAD_LIMIT)),
            "Drone":  ({Cap.AIR},              np.array([10., 10.]), (L.TEMP_LIMIT, L.RAD_LIMIT)),
            "Boat":   ({Cap.WATER},            np.array([0., 0.]),   (9999., 9999.)),
            "Rover":  ({Cap.LAND},             np.array([-2., -2.]), (9999., 9999.)),
        }
        spawn = []
        for t, n in counts.items(): spawn += [t] * n
        random.shuffle(spawn)
        W, H = L.GRID_W, L.GRID_H
        clusters = [(W//6,H//6),(W//2,H//6),(5*W//6,H//6),
                    (W//6,H//2),(W//2,H//2),(5*W//6,H//2),
                    (W//6,5*H//6),(W//2,5*H//6),(5*W//6,5*H//6)]
        water_cells = [(x, y) for x in range(W) for y in range(H)
                       if self.world.grid[x][y]["t"] == L.T_WATER]
        self.robots = []
        for i, tname in enumerate(spawn):
            caps, weights, (tlim, rlim) = templates[tname]
            center = clusters[i % len(clusters)]
            if tname == "Boat" and water_cells:
                ns = [c for c in water_cells if not self.radio_shadow[c[0], c[1]]] or water_cells
                sx, sy = random.choice(ns)
            else:
                sx, sy = center
                for _ in range(60):
                    cx = max(1, min(W-2, center[0] + random.randint(-14, 14)))
                    cy = max(1, min(H-2, center[1] + random.randint(-14, 14)))
                    tt = self.world.grid[cx][cy]["t"]
                    if (tt == L.T_FREE or (tt == L.T_STAIRS and L.Capability.STAIRS in caps)) \
                            and not self.radio_shadow[cx, cy]:
                        sx, sy = cx, cy; break
            self.robots.append(L.Robot(f"{tname}{i}", sx, sy, caps, self.world,
                                        self, weights, tlim, rlim))

    def _build_survivors(self):
        W, H = L.GRID_W, L.GRID_H
        free_open = [(x, y) for x in range(W) for y in range(H)
                     if self.world.grid[x][y]["t"] == L.T_FREE]
        stair = [(x, y) for x in range(W) for y in range(H)
                 if self.world.grid[x][y]["t"] == L.T_STAIRS]
        random.shuffle(free_open); random.shuffle(stair)
        n_stair = min(len(stair), int(round(n_survivors * 0.55)))
        chosen = stair[:n_stair] + free_open[:max(0, n_survivors - n_stair)]
        self.survivors = chosen[:n_survivors]

    L.FleetSim._build_robots = _build_robots
    L.FleetSim._build_survivors = _build_survivors
    return L.FleetSim()

# ── Run one (N, seed) ────────────────────────────────────────────────────────
def run_one(n_robots, seed, steps, fixed_grid):
    grid = grid_for(n_robots, fixed_grid)
    n_surv = max(5, int(round(grid * grid * SURV_PER_CELL)))
    random.seed(seed); np.random.seed(seed)
    sim = make_sim(n_robots, grid, n_surv)
    actual_n = len(sim.robots); actual_surv = len(sim.survivors)

    # phase timers
    ph = defaultdict(float)
    def timed(method, key):
        def w(*a, **k):
            t0 = time.perf_counter()
            try: return method(*a, **k)
            finally: ph[key] += time.perf_counter() - t0
        return w
    def timed_phase(method, key, phase_tag):
        # Like timed(), but also flags _CURRENT_PHASE for the duration of the
        # call so any AStar.search invoked underneath is attributed correctly
        # in _ASTAR_PHASE. Restores the previous tag on exit (these calls are
        # not reentrant/nested in this sim, but this is defensive regardless).
        def w(*a, **k):
            prev = _CURRENT_PHASE['p']; _CURRENT_PHASE['p'] = phase_tag
            t0 = time.perf_counter()
            try: return method(*a, **k)
            finally:
                ph[key] += time.perf_counter() - t0
                _CURRENT_PHASE['p'] = prev
        return w
    if hasattr(sim, '_pg_best_response_roles'):
        sim._pg_best_response_roles = timed(sim._pg_best_response_roles, 'role')
    if hasattr(sim, '_assign_zones_cbba'):
        sim._assign_zones_cbba = timed(sim._assign_zones_cbba, 'alloc')
    if hasattr(sim, '_active_relay_coverage_union'):
        sim._active_relay_coverage_union = timed(sim._active_relay_coverage_union, 'relay_cover')
    # explorer_move / relay_move: whole-method time for the two call sites
    # that own every AStar.search call in the framework (goal selection +
    # replanning for explorers; relay border pathing for relays). This is
    # what used to be invisible inside the unlabelled 'other' residual.
    if hasattr(sim, '_move_robot'):
        sim._move_robot = timed_phase(sim._move_robot, 'explorer_move', 'explorer')
    if hasattr(sim, '_move_relay'):
        sim._move_relay = timed_phase(sim._move_relay, 'relay_move', 'relay')

    step_ms, ground_ms, amax_ms, asum_ms, dist_ms = [], [], [], [], []
    astar_explorer_ms, astar_relay_ms = [], []
    t_first = t_half = t_all = -1
    peak_relays = 0
    t_wall = time.time()
    Role = getattr(L, 'Role', None)

    for step in range(1, steps + 1):
        _ASTAR['sum'] = 0.0; _ASTAR['max'] = 0.0; _ASTAR['n'] = 0
        _ASTAR_PHASE['explorer'] = 0.0; _ASTAR_PHASE['relay'] = 0.0; _ASTAR_PHASE['other'] = 0.0
        t0 = time.perf_counter()
        alive = sim.step()
        dt_ms = (time.perf_counter() - t0) * 1000.0
        asum = _ASTAR['sum']; amax = _ASTAR['max']
        ground = max(0.0, dt_ms - asum)
        step_ms.append(dt_ms); asum_ms.append(asum); amax_ms.append(amax)
        ground_ms.append(ground); dist_ms.append(ground + amax)
        astar_explorer_ms.append(_ASTAR_PHASE['explorer'])
        astar_relay_ms.append(_ASTAR_PHASE['relay'])

        nf = len(sim.found)
        if nf >= 1 and t_first < 0: t_first = step
        if nf >= actual_surv // 2 and t_half < 0: t_half = step
        if nf >= actual_surv and t_all < 0: t_all = step
        if Role is not None:
            nr = sum(1 for r in sim.robots if r.active and r.role == Role.RELAY)
            peak_relays = max(peak_relays, nr)
        if not alive: break

    wall = time.time() - t_wall
    steps_run = sim.timestep
    cov = float(np.mean(sim.union_belief != 0)) * 100.0
    found = len(sim.found)
    deaths = sum(1 for r in sim.robots if not r.active)
    hazard = sum(1 for r in sim.robots if getattr(r, 'hazard_killed', False))
    battery = sum(1 for _, rs in getattr(sim, 'dead_robots', []) if 'batt' in (rs or '').lower())

    def stats(a):
        if not a: return dict(mean=0, p50=0, p90=0, max=0)
        A = np.array(a)
        return dict(mean=float(A.mean()), p50=float(np.percentile(A,50)),
                     p90=float(np.percentile(A,90)), max=float(A.max()))
    S = {k: stats(v) for k, v in
         dict(step=step_ms, ground=ground_ms, astar_sum=asum_ms,
              astar_max=amax_ms, dist=dist_ms,
              astar_explorer=astar_explorer_ms, astar_relay=astar_relay_ms).items()}

    # phase totals → ms/step
    role_ms  = ph['role']  * 1000.0 / max(1, steps_run)
    alloc_ms = ph['alloc'] * 1000.0 / max(1, steps_run)
    relay_cover_ms = ph['relay_cover'] * 1000.0 / max(1, steps_run)
    explorer_move_ms = ph['explorer_move'] * 1000.0 / max(1, steps_run)
    relay_move_ms    = ph['relay_move']    * 1000.0 / max(1, steps_run)
    # Genuine residual: whatever step_total accounts for that isn't one of the
    # five measured phases (comms mail delivery, survivor-detection sweep,
    # snapshot bookkeeping, etc). Should be small; if it isn't, that's a sign
    # something expensive is happening outside all five wrapped methods.
    other_ms = max(0.0, S['step']['mean'] - role_ms - alloc_ms - relay_cover_ms
                   - explorer_move_ms - relay_move_ms)

    row = dict(
        n_requested=n_robots, n_actual=actual_n, grid=grid,
        cells_per_robot=round(grid*grid/max(1,actual_n), 1),
        seed=seed, steps_run=steps_run, survivors=actual_surv,
        final_cov=round(cov,2), final_found=found,
        found_frac=round(found/max(1,actual_surv),3),
        deaths=deaths, hazard_deaths=hazard, battery_deaths=battery,
        peak_relays=peak_relays,
        t_first_found=t_first, t_half_found=t_half, t_all_found=t_all,
        completed=int(t_all > 0),
        wall_s=round(wall,1),
        step_mean_ms=round(S['step']['mean'],3),  step_p50_ms=round(S['step']['p50'],3),
        step_p90_ms=round(S['step']['p90'],3),    step_max_ms=round(S['step']['max'],3),
        ground_mean_ms=round(S['ground']['mean'],3), ground_p90_ms=round(S['ground']['p90'],3),
        astar_sum_mean_ms=round(S['astar_sum']['mean'],3),
        astar_max_mean_ms=round(S['astar_max']['mean'],3),
        astar_max_p90_ms=round(S['astar_max']['p90'],3),
        dist_mean_ms=round(S['dist']['mean'],3), dist_p90_ms=round(S['dist']['p90'],3),
        role_ms=round(role_ms,3), alloc_ms=round(alloc_ms,3), relay_cover_ms=round(relay_cover_ms,3),
        # ── Mutually-exclusive method-level breakdown (sums to ≈ step_mean_ms) ──
        # explorer_move / relay_move are whole-method time for _move_robot and
        # _move_relay — the two call sites that own EVERY AStar.search call in
        # the framework. astar_explorer_ms / astar_relay_ms are the A*-only
        # SUBSET of each (so explorer_move_ms − astar_explorer_ms is that
        # method's own non-A* bookkeeping: goal scoring, hysteresis, etc).
        # 'alloc' (CBBA) and 'role' (potential game) call NO A* at all — this
        # breakdown is what makes that visible instead of hiding it in a
        # catch-all 'other'.
        explorer_move_ms=round(explorer_move_ms,3), relay_move_ms=round(relay_move_ms,3),
        astar_explorer_ms=round(S['astar_explorer']['mean'],3),
        astar_relay_ms=round(S['astar_relay']['mean'],3),
        other_ms=round(other_ms,3),
        # per-robot normalisations (reveal O(N) vs O(1) phase scaling)
        step_ms_per_robot=round(S['step']['mean']/max(1,actual_n),4),
        ground_ms_per_robot=round(S['ground']['mean']/max(1,actual_n),4),
        role_ms_per_robot=round(role_ms/max(1,actual_n),4),
        alloc_ms_per_robot=round(alloc_ms/max(1,actual_n),4),
        relay_cover_ms_per_robot=round(relay_cover_ms/max(1,actual_n),4),
        explorer_move_ms_per_robot=round(explorer_move_ms/max(1,actual_n),4),
        relay_move_ms_per_robot=round(relay_move_ms/max(1,actual_n),4),
    )
    return row

# ── Excel writer ─────────────────────────────────────────────────────────────
def write_xlsx(rows, agg, cfg, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference

    HEAD = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    HFILL = PatternFill('solid', fgColor='2F5496')
    NORM = Font(name='Arial', size=10)
    CEN = Alignment(horizontal='center'); THIN = Side(style='thin', color='D9D9D9')
    BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def sheet_from(ws, headers, data_rows):
        ws.append(headers)
        for c in range(1, len(headers)+1):
            cell = ws.cell(1, c); cell.font = HEAD; cell.fill = HFILL
            cell.alignment = CEN; cell.border = BORD
        for r in data_rows:
            ws.append([r.get(h) for h in headers])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = NORM; cell.border = BORD
                if isinstance(cell.value, float): cell.alignment = CEN
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(16, max(9, w+2))
        ws.freeze_panes = 'A2'

    wb = Workbook()

    # 1) Raw per-(N,seed)
    raw_headers = list(rows[0].keys())
    sheet_from(wb.active, raw_headers, rows); wb.active.title = 'raw_by_seed'

    # 2) Aggregated by N (means across seeds) with Excel formulas linking to raw
    ws = wb.create_sheet('agg_by_N')
    agg_headers = list(agg[0].keys())
    sheet_from(ws, agg_headers, agg)

    # 3) Phase-scaling sheet: mutually-exclusive method-level breakdown, plus
    #    ratios computed with Excel formulas. explorer_move/relay_move own
    #    EVERY AStar.search call in the framework; role (potential game) and
    #    alloc (CBBA) call none — astar_explorer_ms/astar_relay_ms make that
    #    split explicit instead of leaving A* buried in an unlabelled 'other'.
    ws3 = wb.create_sheet('phase_scaling')
    ph_headers = ['n_actual', 'step_mean_ms', 'role_ms', 'alloc_ms', 'relay_cover_ms',
                  'explorer_move_ms', 'relay_move_ms', 'other_ms',
                  'astar_explorer_ms', 'astar_relay_ms',
                  'ground_mean_ms', 'astar_sum_mean_ms', 'astar_max_mean_ms',
                  'dist_mean_ms']
    sheet_from(ws3, ph_headers, agg)
    # add derived columns via formulas (shares of step_mean_ms, per-robot)
    last = ws3.max_row
    extra = ['astar_total_share_%', 'alloc_share_%', 'role_share_%',
             'ground_share_%', 'ground_per_robot_ms']
    base = len(ph_headers)
    col = {name: chr(ord('A') + i) for i, name in enumerate(ph_headers)}
    for j, name in enumerate(extra):
        c = ws3.cell(1, base+1+j, name); c.font = HEAD; c.fill = HFILL; c.alignment = CEN; c.border = BORD
    for i in range(2, last+1):
        # astar_explorer + astar_relay as % of total step time — the direct
        # answer to "is A* the majority of the cost".
        ws3.cell(i, base+1, f'=100*({col["astar_explorer_ms"]}{i}+{col["astar_relay_ms"]}{i})/{col["step_mean_ms"]}{i}').font = NORM
        ws3.cell(i, base+2, f'=100*{col["alloc_ms"]}{i}/{col["step_mean_ms"]}{i}').font = NORM
        ws3.cell(i, base+3, f'=100*{col["role_ms"]}{i}/{col["step_mean_ms"]}{i}').font = NORM
        ws3.cell(i, base+4, f'=100*{col["ground_mean_ms"]}{i}/{col["step_mean_ms"]}{i}').font = NORM
        ws3.cell(i, base+5, f'={col["ground_mean_ms"]}{i}/{col["n_actual"]}{i}').font = NORM
    for c in ws3.columns:
        ws3.column_dimensions[c[0].column_letter].width = 16

    # 3b) Stacked-area chart: role/alloc/cover/explorer_move/relay_move/other
    #     vs N — the visual answer to "where does the cost go as N grows".
    from openpyxl.chart import AreaChart
    achart = AreaChart(grouping='stacked')
    achart.title = "Framework L — mutually-exclusive step-time breakdown vs fleet size"
    achart.x_axis.title = "robots (N)"; achart.y_axis.title = "ms/step"
    for name in ('role_ms', 'alloc_ms', 'relay_cover_ms', 'explorer_move_ms',
                 'relay_move_ms', 'other_ms'):
        idx = ph_headers.index(name) + 1
        data = Reference(ws3, min_col=idx, min_row=1, max_row=len(agg)+1)
        achart.add_data(data, titles_from_data=True)
    cats3 = Reference(ws3, min_col=ph_headers.index('n_actual')+1,
                       min_row=2, max_row=len(agg)+1)
    achart.set_categories(cats3)
    ws3.add_chart(achart, f'A{last+4}')

    # 4) Chart: ground_mean_ms vs n_actual
    chart = LineChart(); chart.title = "Framework L — coordination cost vs fleet size"
    chart.x_axis.title = "robots (N)"; chart.y_axis.title = "ms/step"
    data = Reference(ws, min_col=agg_headers.index('ground_mean_ms')+1,
                      min_row=1, max_row=len(agg)+1)
    cats = Reference(ws, min_col=agg_headers.index('n_actual')+1,
                      min_row=2, max_row=len(agg)+1)
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    ws.add_chart(chart, f'A{len(agg)+4}')

    # 5) Config sheet
    wsc = wb.create_sheet('config')
    wsc.append(['key','value'])
    wsc.cell(1,1).font = HEAD; wsc.cell(1,1).fill = HFILL
    wsc.cell(1,2).font = HEAD; wsc.cell(1,2).fill = HFILL
    for k, v in cfg.items():
        wsc.append([k, str(v)])
    wsc.column_dimensions['A'].width = 26; wsc.column_dimensions['B'].width = 50

    wb.save(path)

# ── CSV writers (tidy long + wide) ───────────────────────────────────────────
def write_csv(rows, wide_path, long_path):
    headers = list(rows[0].keys())
    with open(wide_path, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=headers); w.writeheader()
        for r in rows: w.writerow(r)
    id_cols = ['n_requested','n_actual','grid','cells_per_robot','seed']
    metric_cols = [h for h in headers if h not in id_cols]
    with open(long_path, 'w', newline='') as f:
        w = csv.writer(f); w.writerow(id_cols + ['metric','value'])
        for r in rows:
            for m in metric_cols:
                w.writerow([r[c] for c in id_cols] + [m, r[m]])

# ── Plot ─────────────────────────────────────────────────────────────────────
def plot_curve(agg, path):
    import matplotlib; matplotlib.use('Agg'); import matplotlib.pyplot as plt
    N = [a['n_actual'] for a in agg]
    fig, ax = plt.subplots(3, 3, figsize=(19, 13))
    ax[0,0].plot(N, [a['ground_mean_ms'] for a in agg], 'o-', label='ground (coord)')
    ax[0,0].plot(N, [a['step_mean_ms'] for a in agg], 's--', label='full step (sim)')
    ax[0,0].plot(N, [a['dist_mean_ms'] for a in agg], '^:', label='dist latency')
    ax[0,0].set_title('Compute vs fleet size'); ax[0,0].set_xlabel('N'); ax[0,0].set_ylabel('ms/step')
    ax[0,0].legend(); ax[0,0].grid(alpha=.3)
    ax[0,1].plot(N, [a['ground_ms_per_robot'] for a in agg], 'o-', color='purple')
    ax[0,1].set_title('Coordination ms per robot (flat = O(N) total)')
    ax[0,1].set_xlabel('N'); ax[0,1].set_ylabel('ms/step/robot'); ax[0,1].grid(alpha=.3)

    # Mutually-exclusive method-level breakdown, stacked — sums to ≈ step_mean_ms.
    # explorer_move/relay_move own EVERY AStar.search call; role (potential
    # game) and alloc (CBBA) call none. This is the direct answer to "how
    # much of the cost is A*, and is it hiding inside alloc" — it isn't; it
    # was hiding in the unlabelled 'other' this breakdown replaces.
    stack_keys = [('role_ms', 'role (PG)', '#ff7f0e'),
                  ('alloc_ms', 'alloc (CBBA)', '#2ca02c'),
                  ('relay_cover_ms', 'relay cover (union)', '#d62728'),
                  ('explorer_move_ms', 'explorer move (incl. A*)', '#1f77b4'),
                  ('relay_move_ms', 'relay move (incl. A*)', '#9467bd'),
                  ('other_ms', 'other (unaccounted)', '#7f7f7f')]
    stacks = [[a[k] for a in agg] for k, _, _ in stack_keys]
    ax[1,0].stackplot(N, *stacks, labels=[lbl for _, lbl, _ in stack_keys],
                       colors=[c for _, _, c in stack_keys], alpha=0.85)
    ax[1,0].set_title('Step-time breakdown (stacked, mutually exclusive)')
    ax[1,0].set_xlabel('N'); ax[1,0].set_ylabel('ms/step')
    ax[1,0].legend(fontsize=7, loc='upper left'); ax[1,0].grid(alpha=.3)

    # Same breakdown, normalised to 100% per N — isolates SHARE evolution from
    # absolute growth. A phase can grow in absolute ms while shrinking as a
    # share (e.g. if another phase grows faster) — this panel shows which.
    tot = [max(1e-9, sum(a[k] for k, _, _ in stack_keys)) for a in agg]
    stacks_pct = [[100*v/t for v, t in zip(s, tot)] for s in stacks]
    ax[2,0].stackplot(N, *stacks_pct, labels=[lbl for _, lbl, _ in stack_keys],
                       colors=[c for _, _, c in stack_keys], alpha=0.85)
    ax[2,0].set_title('Step-time breakdown — SHARE (normalised to 100%)')
    ax[2,0].set_xlabel('N'); ax[2,0].set_ylabel('% of accounted step time')
    ax[2,0].set_ylim(0, 100); ax[2,0].grid(alpha=.3)

    # A* as a % of total step time, split by call site (explorer vs relay),
    # against alloc% and role% for direct comparison.
    astar_total = [a['astar_explorer_ms'] + a['astar_relay_ms'] for a in agg]
    astar_pct   = [100*t/max(1e-9, a['step_mean_ms']) for t, a in zip(astar_total, agg)]
    astar_exp_pct = [100*a['astar_explorer_ms']/max(1e-9, a['step_mean_ms']) for a in agg]
    astar_rly_pct = [100*a['astar_relay_ms']/max(1e-9, a['step_mean_ms']) for a in agg]
    alloc_pct = [100*a['alloc_ms']/max(1e-9, a['step_mean_ms']) for a in agg]
    role_pct  = [100*a['role_ms']/max(1e-9, a['step_mean_ms']) for a in agg]
    ax[0,2].plot(N, astar_pct, 'o-', color='#1f77b4', lw=2.4, label='A* total (explorer+relay)')
    ax[0,2].plot(N, astar_exp_pct, 'o--', color='#1f77b4', alpha=0.5, label='  A* — explorer only')
    ax[0,2].plot(N, astar_rly_pct, 'o--', color='#9467bd', alpha=0.5, label='  A* — relay only')
    ax[0,2].plot(N, alloc_pct, 's-', color='#2ca02c', label='alloc (CBBA)')
    ax[0,2].plot(N, role_pct, '^-', color='#ff7f0e', label='role (PG)')
    ax[0,2].set_title('Share of step time by phase (%)')
    ax[0,2].set_xlabel('N'); ax[0,2].set_ylabel('% of step_mean_ms')
    ax[0,2].legend(fontsize=7); ax[0,2].grid(alpha=.3); ax[0,2].set_ylim(0, 100)

    ax[1,1].plot(N, [a['final_found_frac_mean']*100 for a in agg], 'o-', color='green', label='found %')
    ax[1,1].plot(N, [a['final_cov'] for a in agg], 's--', color='teal', label='map coverage %')
    ax[1,1].set_title('Mission outcome vs fleet size'); ax[1,1].set_xlabel('N')
    ax[1,1].set_ylabel('%'); ax[1,1].legend(); ax[1,1].grid(alpha=.3); ax[1,1].set_ylim(0,105)

    ax[1,2].plot(N, [a['explorer_move_ms_per_robot'] for a in agg], 'o-',
                 color='#1f77b4', label='explorer move / robot')
    ax[1,2].plot(N, [a['relay_move_ms_per_robot'] for a in agg], 'o-',
                 color='#9467bd', label='relay move / robot')
    ax[1,2].set_title('A*-owning phases, per robot (flat = O(N) total)')
    ax[1,2].set_xlabel('N'); ax[1,2].set_ylabel('ms/step/robot')
    ax[1,2].legend(fontsize=8); ax[1,2].grid(alpha=.3)

    # Log-log complexity fit: slope of log(metric) vs log(N) is the empirical
    # exponent — 1.0 is linear, >1 is superlinear. Needs >=2 positive points;
    # silently skipped otherwise (e.g. --quick with a single size).
    def fit_slope(xs, ys):
        pts = [(x, y) for x, y in zip(xs, ys) if x > 0 and y > 0]
        if len(pts) < 2:
            return None
        lx = np.log([p[0] for p in pts]); ly = np.log([p[1] for p in pts])
        b, _ = np.polyfit(lx, ly, 1)
        return b
    loglog_keys = [('step_mean_ms', 'step (total)', '#000000'),
                   ('explorer_move_ms', 'explorer move', '#1f77b4'),
                   ('alloc_ms', 'alloc (CBBA)', '#2ca02c'),
                   ('relay_cover_ms', 'relay cover', '#d62728')]
    for k, lbl, c in loglog_keys:
        ys = [a[k] for a in agg]
        s = fit_slope(N, ys)
        tag = f'{lbl}  (N^{s:.2f})' if s is not None else lbl
        ax[2,1].loglog(N, ys, 'o-', color=c, label=tag)
    ax[2,1].set_title('Log-log fit — empirical scaling exponent')
    ax[2,1].set_xlabel('N (log)'); ax[2,1].set_ylabel('ms/step (log)')
    ax[2,1].legend(fontsize=7); ax[2,1].grid(alpha=.3, which='both')

    # Central (ground station, non-parallelisable) vs the slowest single
    # robot's own onboard plan (parallelisable, one CPU per robot in a real
    # deployment) — the two numbers that actually determine feasibility of
    # running this coordination scheme centrally vs distributed.
    ax[2,2].plot(N, [a['ground_mean_ms'] for a in agg], 'o-', color='#e07b00',
                 label='ground station (central, non-parallel)')
    ax[2,2].plot(N, [a['astar_max_mean_ms'] for a in agg], 's-', color='#1f77b4',
                 label="slowest robot's own plan (parallel)")
    ax[2,2].set_title('Central vs. parallel latency bottleneck')
    ax[2,2].set_xlabel('N'); ax[2,2].set_ylabel('ms/step')
    ax[2,2].legend(fontsize=8); ax[2,2].grid(alpha=.3)

    fig.suptitle('Framework L — fleet-size scaling', fontsize=13, fontweight='bold')
    fig.tight_layout(); fig.savefig(path, dpi=120, bbox_inches='tight'); plt.close(fig)

def _parse_seeds(spec):
    """--seeds accepts either a COUNT ('10' -> seeds 0..9) or an EXPLICIT
    comma list ('1,2,7,42' -> exactly those seeds), same style as --sizes.
    A bare count is what most runs want; an explicit list lets you add specific
    seeds to an existing dataset without renumbering or rerunning 0..N-1."""
    spec = str(spec).strip()
    if ',' in spec:
        return [int(s) for s in spec.split(',') if s.strip() != '']
    return list(range(int(spec)))

# ── Aggregate across seeds ───────────────────────────────────────────────────
def aggregate(rows):
    by_n = defaultdict(list)
    for r in rows: by_n[r['n_requested']].append(r)
    agg = []
    for n in sorted(by_n):
        g = by_n[n]
        def mean(k): return round(float(np.mean([x[k] for x in g])), 3)
        agg.append(dict(
            n_requested=n, n_actual=g[0]['n_actual'], grid=g[0]['grid'],
            cells_per_robot=g[0]['cells_per_robot'], n_seeds=len(g),
            final_cov=mean('final_cov'),
            final_found_frac_mean=mean('found_frac'),
            completed_frac=round(float(np.mean([x['completed'] for x in g])),3),
            deaths=mean('deaths'), hazard_deaths=mean('hazard_deaths'),
            battery_deaths=mean('battery_deaths'), peak_relays=mean('peak_relays'),
            t_all_found=mean('t_all_found'),
            step_mean_ms=mean('step_mean_ms'), step_p90_ms=mean('step_p90_ms'),
            ground_mean_ms=mean('ground_mean_ms'), ground_p90_ms=mean('ground_p90_ms'),
            astar_sum_mean_ms=mean('astar_sum_mean_ms'),
            astar_max_mean_ms=mean('astar_max_mean_ms'),
            dist_mean_ms=mean('dist_mean_ms'), dist_p90_ms=mean('dist_p90_ms'),
            role_ms=mean('role_ms'), alloc_ms=mean('alloc_ms'), relay_cover_ms=mean('relay_cover_ms'),
            explorer_move_ms=mean('explorer_move_ms'), relay_move_ms=mean('relay_move_ms'),
            astar_explorer_ms=mean('astar_explorer_ms'), astar_relay_ms=mean('astar_relay_ms'),
            other_ms=mean('other_ms'),
            ground_ms_per_robot=mean('ground_ms_per_robot'),
            step_ms_per_robot=mean('step_ms_per_robot'),
            explorer_move_ms_per_robot=mean('explorer_move_ms_per_robot'),
            relay_move_ms_per_robot=mean('relay_move_ms_per_robot'),
        ))
    return agg

# ── Empirical scaling exponents + reviewer report ────────────────────────────
def fit_exponents(agg):
    """Log-log slope of each phase vs N — the empirical scaling exponent
    (1.0 = linear, >1 superlinear). Needs >=2 sizes with positive values."""
    N = [a['n_actual'] for a in agg]
    keys = ['step_mean_ms', 'ground_mean_ms', 'astar_max_mean_ms',
            'explorer_move_ms', 'relay_move_ms', 'alloc_ms', 'role_ms',
            'relay_cover_ms', 'step_ms_per_robot', 'ground_ms_per_robot']
    out = []
    for k in keys:
        pts = [(n, a[k]) for n, a in zip(N, agg) if n > 0 and a.get(k, 0) > 0]
        if len(pts) < 2:
            out.append((k, None, len(pts))); continue
        lx = np.log([p[0] for p in pts]); ly = np.log([p[1] for p in pts])
        b, _ = np.polyfit(lx, ly, 1)
        out.append((k, round(float(b), 3), len(pts)))
    return out


def reviewer_report(agg, exps, cfg):
    """The scaling metrics to put in front of reviewers, printed and framed.
    The claim structure that survives review:
      1. CENTRAL coordination exponent (ground_mean_ms) — the part one ground
         station must do; superlinear here is the honest cost to report.
      2. PARALLEL per-robot exponent (astar_max_mean_ms — slowest single
         robot's own plan) — near-linear/flat means onboard planning
         distributes, so absolute step-ms is a simulator artifact, not a
         deployment property.
      3. WHICH phase is superlinear (alloc vs explorer_move) — explorer_move
         growth partly reflects sqrt(N) path-length growth under density-held
         weak scaling (geometry, not coordination); alloc growth is the
         coordination story, and the Layer-0 cluster cap is its knob
         (--cluster-cap 0 to test).
      4. QUALITY FLAT check — found%/coverage stable across N means the
         coordination does not degrade with scale, only its cost grows."""
    e = {k: v for k, v, _ in exps}
    def fmt(k):
        v = e.get(k)
        return f'N^{v:.2f}' if v is not None else 'n/a (<2 sizes)'
    ff = [a['final_found_frac_mean'] for a in agg]
    cv = [a['final_cov'] for a in agg]
    print("\n  REVIEWER HEADLINE — scaling exponents (log-log fits across sizes)")
    print(f"    central coordination (ground station):   {fmt('ground_mean_ms')}")
    print(f"    parallel per-robot plan (slowest A*):    {fmt('astar_max_mean_ms')}")
    print(f"    per-robot coordination (flat => O(N)):   {fmt('ground_ms_per_robot')}")
    print(f"    alloc / HRBA auction:                    {fmt('alloc_ms')}"
          f"   <- the superlinear suspect; knob = --cluster-cap")
    print(f"    explorer movement (incl. A*):            {fmt('explorer_move_ms')}"
          f"   (part is sqrt(N) path geometry under weak scaling)")
    print(f"    role / potential game:                   {fmt('role_ms')}")
    print(f"    quality across N: found_frac {min(ff):.2f}-{max(ff):.2f}, "
          f"coverage {min(cv):.1f}-{max(cv):.1f}%  (flat = no degradation)")
    if cfg.get('cluster_cap') is not None:
        print(f"    cluster cap this run: {cfg['cluster_cap']}"
              f" (compare against a default-cap run)")
    if len(agg) < 4:
        print("    (fewer than 4 sizes — exponents are illustrative, not citable)")


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser()
    p.add_argument('--sizes', type=str, default='10,20,30,40,50,60,70,80,90,100,1000')
    p.add_argument('--seeds', type=str, default='10',
                    help="Either a COUNT (e.g. '10' -> seeds 0..9) or an "
                         "explicit comma list (e.g. '1,2,7,42') to add "
                         "specific seeds without renumbering. Default: 10.")
    p.add_argument('--steps', type=int, default=1500)
    p.add_argument('--fixed-grid', action='store_true',
                    help='hold map at 200×200 (density rises with N) instead of density-held growth')
    p.add_argument('--max-cells', type=int, default=360000,
                    help='skip N whose density-held grid exceeds this many cells (sim memory ceiling; ~600x600). Default 360000.')
    p.add_argument('--quick', action='store_true')
    p.add_argument('--out', type=str, default='scaling_L_out')
    p.add_argument('--cluster-cap', type=int, default=None,
                    help="override the framework's HRBA Layer-0 cluster cap "
                         "(shipped cap = 8, which saturates at ~64 robots — "
                         "exactly where alloc cost bends superlinear). "
                         "0 = uncapped (k = sqrt(R)); n>=1 = cap at n. "
                         "Applied to a temp copy; the framework file is never "
                         "modified. Compare runs at different caps to test "
                         "whether the hierarchy is the scaling mechanism.")
    a = p.parse_args()
    if a.cluster_cap is not None:
        global L
        L = _load_framework(a.cluster_cap)
        _wrap_astar(L.AStar)      # re-instrument the fresh module's A*
        cap_desc = 'UNCAPPED (k=sqrt(R))' if a.cluster_cap == 0 else f'cap={a.cluster_cap}'
        print(f"  [HRBA Layer-0 cluster cap override: {cap_desc}]")
    if a.quick:
        a.sizes = '10,20'; a.seeds = '1'; a.steps = 120
    sizes = [int(s) for s in a.sizes.split(',')]
    seed_list = _parse_seeds(a.seeds)
    os.makedirs(a.out, exist_ok=True)

    print(f"Framework L scaling | sizes={sizes} | seeds={seed_list} | steps={a.steps} "
          f"| {'FIXED 200×200' if a.fixed_grid else 'density-held grid'}")
    rows = []
    skipped = []
    t0 = time.time()
    for n in sizes:
        g = grid_for(n, a.fixed_grid)
        # Guard rail: very large grids exhaust memory in this single-process,
        # dense-array simulator. Flag rather than crash — the ceiling is a
        # SIMULATION-IMPLEMENTATION limit (full-grid arrays × per-robot A*),
        # not a Framework-L coordination limit, and is reported as such.
        est_cells = g * g
        if est_cells > a.max_cells:
            print(f"  N={n:<5} grid={g}×{g} ({est_cells:,} cells) SKIPPED "
                  f"— exceeds --max-cells={a.max_cells:,} (sim memory ceiling, "
                  f"not a framework limit)")
            skipped.append(dict(n_requested=n, grid=g, cells=est_cells,
                                 reason='exceeds_sim_memory_ceiling'))
            continue
        for s in seed_list:
            print(f"  N={n:<5} grid={g}×{g} seed={s} ... ", end='', flush=True)
            try:
                r = run_one(n, s, a.steps, a.fixed_grid)
                print(f"cov={r['final_cov']:5.1f}% found={r['final_found']}/{r['survivors']} "
                      f"ground={r['ground_mean_ms']:.1f}ms step={r['step_mean_ms']:.1f}ms "
                      f"dist={r['dist_mean_ms']:.1f}ms")
                rows.append(r)
            except MemoryError:
                print("MemoryError — sim ceiling reached, flagged")
                skipped.append(dict(n_requested=n, grid=g, cells=est_cells,
                                     reason='MemoryError_at_runtime')); break
            except Exception as e:
                import traceback; print("FAIL:", e); traceback.print_exc()
    print(f"total {(time.time()-t0)/60:.1f} min")

    if not rows:
        print("no rows — aborting"); return
    agg = aggregate(rows)
    cfg = dict(framework='L (belief-safe K)', sizes=sizes, seeds=seed_list, steps=a.steps,
               cluster_cap=a.cluster_cap,
               grid_policy=('fixed 200x200' if a.fixed_grid else 'density-held (~800 cells/robot)'),
               type_mix='Legged26/Drone34/Boat16/Rover24 %',
               survivor_density=f'{SURV_PER_CELL:.6f} per cell',
               pythonhashseed=os.environ.get('PYTHONHASHSEED', 'unset'),
               skipped_sizes=';'.join(f"N={x['n_requested']}(grid={x['grid']},{x['reason']})" for x in skipped) or 'none',
               generated=time.strftime('%Y-%m-%d %H:%M:%S'))
    write_csv(rows, os.path.join(a.out, 'scaling_L_wide.csv'),
                    os.path.join(a.out, 'scaling_L_long.csv'))
    written = ['scaling_L_long.csv', 'scaling_L_wide.csv']
    exps = fit_exponents(agg)
    with open(os.path.join(a.out, 'scaling_L_exponents.csv'), 'w', newline='') as f:
        w = csv.writer(f); w.writerow(['metric', 'exponent', 'n_sizes'])
        for k, v, n in exps:
            w.writerow([k, '' if v is None else v, n])
    written.append('scaling_L_exponents.csv')
    reviewer_report(agg, exps, cfg)
    try:
        write_xlsx(rows, agg, cfg, os.path.join(a.out, 'scaling_L_summary.xlsx'))
        written.insert(0, 'scaling_L_summary.xlsx')
    except ImportError:
        # openpyxl not installed on this machine — write the aggregated summary
        # as plain CSV instead so nothing is lost, and say how to get the xlsx.
        import csv as _csv
        p = os.path.join(a.out, 'scaling_L_summary.csv')
        with open(p, 'w', newline='') as f:
            w = _csv.writer(f)
            if agg:
                keys = list(agg[0].keys())
                w.writerow(keys)
                for row in agg:
                    w.writerow([row.get(k, '') for k in keys])
            w.writerow([])
            w.writerow(['config'])
            for k, v in cfg.items():
                w.writerow([k, v])
        written.insert(0, 'scaling_L_summary.csv')
        print("  [openpyxl not installed — summary written as scaling_L_summary.csv "
              "instead; `pip install openpyxl` to get the formatted workbook]")
    except Exception as e:
        print("xlsx write failed:", e)
    try:
        plot_curve(agg, os.path.join(a.out, 'scaling_L_curve.png'))
        written.append('scaling_L_curve.png')
    except Exception as e:
        print("plot failed:", e)
    print(f"\nwrote → {a.out}/  ({', '.join(written)})")

if __name__ == '__main__':
    main()